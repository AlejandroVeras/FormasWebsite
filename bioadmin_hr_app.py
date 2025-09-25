#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aplicación BioAdmin HR - Sistema de Procesamiento de Reportes de RRHH
===============================================================

Autor: Sistema FORMAS HR
Versión: 2.0
Fecha: 2024

Descripción:
Esta aplicación procesa archivos de BioAdmin y reportes de permisos para generar
reportes mensuales profesionales en formato Excel con análisis detallado de
horas trabajadas por empleado.

Características:
- Procesamiento de archivos BioAdmin (.csv, .txt, .xlsx)
- Análisis de reportes de permisos
- Agrupación de datos por mes
- Generación de reportes Excel profesionales
- Interfaz completamente en español
- Manejo robusto de fechas en español
- Validación exhaustiva de archivos
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys
import re
import locale
import warnings
from typing import Dict, List, Tuple, Optional, Union
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.cell.cell import MergedCell
# Import GUI components only if available
GUI_AVAILABLE = False
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    from tkinter.scrolledtext import ScrolledText
    GUI_AVAILABLE = True
except ImportError:
    print("⚠️ Advertencia: tkinter no está disponible. Solo funcionará el modo consola.")
    # Create dummy classes for when GUI is not available
    class tk:
        class Tk: pass
    class ttk:
        class Frame: pass
        class Label: pass
        class Button: pass
        class LabelFrame: pass
        class Progressbar: pass
    class ScrolledText: pass
    
    # Mock functions
    def filedialog(): pass
    def messagebox(): pass

# Suprimir advertencias no críticas
warnings.filterwarnings('ignore', category=UserWarning)

class BioadminHRProcessor:
    """
    Procesador principal para archivos BioAdmin y reportes de RRHH
    """
    
    # Mapeo de meses en español
    MESES_ESPANOL = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
        'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4,
        'may': 5, 'jun': 6, 'jul': 7, 'ago': 8,
        'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
    }
    
    # Colores para reportes Excel
    COLORES_EXCEL = {
        'encabezado': 'FF2F75B5',
        'total': 'FFFFE066',
        'subtotal': 'FFE7E6E6',
        'alerta': 'FFFF9999',
        'exito': 'FF99FF99',
        'info': 'FF99CCFF'
    }
    
    def __init__(self):
        """Inicializa el procesador con configuraciones por defecto"""
        self.datos_bioadmin = None
        self.datos_permisos = None
        self.resumen_mensual = None
        self.errores = []
        self.configurar_locale()
        
    def configurar_locale(self):
        """Configura el locale para manejo correcto de fechas en español"""
        try:
            # Intentar configurar locale español
            for loc in ['es_ES.UTF-8', 'es_ES', 'Spanish_Spain.1252', 'es']:
                try:
                    locale.setlocale(locale.LC_TIME, loc)
                    break
                except locale.Error:
                    continue
        except Exception as e:
            print(f"⚠️  Advertencia: No se pudo configurar locale español: {e}")
    
    def convertir_fecha_espanol(self, fecha_str: str) -> Optional[datetime]:
        """
        Convierte una fecha en español a objeto datetime
        
        Args:
            fecha_str: Fecha en formato string (español)
            
        Returns:
            datetime: Objeto datetime o None si no se puede convertir
        """
        if not fecha_str or pd.isna(fecha_str):
            return None
            
        fecha_str = str(fecha_str).strip().lower()
        
        # Patrones de fecha comunes
        patrones = [
            r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})',  # 15 de enero de 2024
            r'(\d{1,2})/(\w+)/(\d{4})',                 # 15/enero/2024
            r'(\d{1,2})-(\w+)-(\d{4})',                 # 15-enero-2024
            r'(\d{4})-(\d{1,2})-(\d{1,2})',            # 2024-01-15
            r'(\d{1,2})/(\d{1,2})/(\d{4})',            # 15/01/2024
            r'(\d{1,2})-(\d{1,2})-(\d{4})',            # 15-01-2024
        ]
        
        for patron in patrones:
            match = re.match(patron, fecha_str)
            if match:
                grupos = match.groups()
                try:
                    if len(grupos) == 3:
                        if grupos[1].isdigit():  # Formato numérico
                            return datetime(int(grupos[2]), int(grupos[1]), int(grupos[0]))
                        else:  # Formato con mes en español
                            mes_str = grupos[1].lower()
                            if mes_str in self.MESES_ESPANOL:
                                mes_num = self.MESES_ESPANOL[mes_str]
                                return datetime(int(grupos[2]), mes_num, int(grupos[0]))
                except ValueError:
                    continue
                    
        # Intentar conversión directa con pandas
        try:
            return pd.to_datetime(fecha_str, errors='coerce', dayfirst=True)
        except:
            return None
    
    def leer_archivo_bioadmin(self, ruta_archivo: str) -> bool:
        """
        Lee y procesa archivos de BioAdmin
        
        Args:
            ruta_archivo: Ruta al archivo de BioAdmin
            
        Returns:
            bool: True si se procesó correctamente
        """
        try:
            if not os.path.exists(ruta_archivo):
                self.errores.append(f"❌ El archivo no existe: {ruta_archivo}")
                return False
                
            extension = os.path.splitext(ruta_archivo)[1].lower()
            
            # Leer según el tipo de archivo
            if extension == '.csv':
                # Intentar diferentes encodings
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(ruta_archivo, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    raise Exception("No se pudo leer el archivo con ningún encoding")
                    
            elif extension in ['.xlsx', '.xls']:
                df = pd.read_excel(ruta_archivo)
            elif extension == '.txt':
                # Intentar como CSV con diferentes separadores
                separadores = ['\t', ',', ';', '|']
                for sep in separadores:
                    try:
                        df = pd.read_csv(ruta_archivo, sep=sep, encoding='utf-8')
                        if len(df.columns) > 1:
                            break
                    except:
                        continue
                else:
                    df = pd.read_csv(ruta_archivo, encoding='utf-8')
            else:
                self.errores.append(f"❌ Formato de archivo no soportado: {extension}")
                return False
            
            # Validar estructura del archivo
            if df.empty:
                self.errores.append("❌ El archivo está vacío")
                return False
                
            # Normalizar nombres de columnas
            df.columns = [self.normalizar_nombre_columna(col) for col in df.columns]
            
            # Buscar columnas esenciales
            columnas_requeridas = ['empleado', 'fecha', 'hora_entrada', 'hora_salida']
            columnas_encontradas = []
            
            for req in columnas_requeridas:
                for col in df.columns:
                    if req in col.lower() or any(palabra in col.lower() 
                                               for palabra in self.obtener_sinonimos_columna(req)):
                        columnas_encontradas.append(col)
                        break
            
            if len(columnas_encontradas) < len(columnas_requeridas):
                self.errores.append(f"⚠️  Columnas faltantes. Encontradas: {columnas_encontradas}")
            
            self.datos_bioadmin = df
            print(f"✅ Archivo BioAdmin cargado exitosamente: {len(df)} registros")
            return True
            
        except Exception as e:
            self.errores.append(f"❌ Error al leer archivo BioAdmin: {str(e)}")
            return False
    
    def normalizar_nombre_columna(self, nombre: str) -> str:
        """Normaliza nombres de columnas eliminando caracteres especiales"""
        nombre = str(nombre).lower().strip()
        nombre = re.sub(r'[^\w\s]', '', nombre)
        nombre = re.sub(r'\s+', '_', nombre)
        return nombre
    
    def obtener_sinonimos_columna(self, columna: str) -> List[str]:
        """Obtiene sinónimos para nombres de columnas comunes"""
        sinonimos = {
            'empleado': ['empleado', 'trabajador', 'personal', 'nombre', 'usuario', 'id'],
            'fecha': ['fecha', 'date', 'dia', 'cuando'],
            'hora_entrada': ['entrada', 'inicio', 'llegada', 'start', 'in'],
            'hora_salida': ['salida', 'fin', 'final', 'end', 'out']
        }
        return sinonimos.get(columna, [])
    
    def procesar_horas_trabajadas(self) -> bool:
        """
        Procesa las horas trabajadas calculando totales diarios y mensuales
        
        Returns:
            bool: True si se procesó correctamente
        """
        try:
            if self.datos_bioadmin is None:
                self.errores.append("❌ No hay datos de BioAdmin cargados")
                return False
            
            df = self.datos_bioadmin.copy()
            
            # Identificar columnas de fecha y hora
            columna_fecha = self.encontrar_columna(df, ['fecha', 'date', 'dia'])
            columna_empleado = self.encontrar_columna(df, ['empleado', 'nombre', 'usuario'])
            columna_entrada = self.encontrar_columna(df, ['entrada', 'inicio', 'start'])
            columna_salida = self.encontrar_columna(df, ['salida', 'fin', 'end'])
            
            if not all([columna_fecha, columna_empleado]):
                self.errores.append("❌ No se encontraron las columnas básicas necesarias")
                return False
            
            # Procesar fechas
            df['fecha_procesada'] = df[columna_fecha].apply(self.convertir_fecha_espanol)
            df = df.dropna(subset=['fecha_procesada'])
            
            # Extraer información temporal
            df['año'] = df['fecha_procesada'].dt.year
            df['mes'] = df['fecha_procesada'].dt.month
            df['nombre_mes'] = df['fecha_procesada'].dt.month.map({
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            })
            df['dia_semana'] = df['fecha_procesada'].dt.day_name()
            
            # Calcular horas trabajadas si hay información de entrada/salida
            if columna_entrada and columna_salida:
                df['horas_trabajadas'] = self.calcular_horas_diarias(
                    df[columna_entrada], df[columna_salida]
                )
            else:
                # Asumir 8 horas por día por defecto
                df['horas_trabajadas'] = 8.0
                print("⚠️  No se encontraron columnas de hora, asumiendo 8 horas por día")
            
            # Crear resumen mensual
            self.resumen_mensual = df.groupby([
                columna_empleado, 'año', 'mes', 'nombre_mes'
            ]).agg({
                'horas_trabajadas': ['sum', 'mean', 'count'],
                'fecha_procesada': ['min', 'max']
            }).round(2)
            
            # Aplanar columnas multinivel
            self.resumen_mensual.columns = [
                'total_horas', 'promedio_diario', 'dias_trabajados', 
                'fecha_inicio', 'fecha_fin'
            ]
            self.resumen_mensual = self.resumen_mensual.reset_index()
            
            print(f"✅ Procesamiento completado: {len(self.resumen_mensual)} registros mensuales")
            return True
            
        except Exception as e:
            self.errores.append(f"❌ Error al procesar horas: {str(e)}")
            return False
    
    def encontrar_columna(self, df: pd.DataFrame, posibles_nombres: List[str]) -> Optional[str]:
        """Encuentra una columna basada en posibles nombres"""
        for posible in posibles_nombres:
            for columna in df.columns:
                if posible.lower() in columna.lower():
                    return columna
        return None
    
    def calcular_horas_diarias(self, entrada_series: pd.Series, salida_series: pd.Series) -> pd.Series:
        """Calcula horas trabajadas entre entrada y salida"""
        horas_calculadas = []
        
        for entrada, salida in zip(entrada_series, salida_series):
            try:
                if pd.isna(entrada) or pd.isna(salida):
                    horas_calculadas.append(0.0)
                    continue
                
                # Convertir a string y limpiar
                entrada_str = str(entrada).strip()
                salida_str = str(salida).strip()
                
                # Parsear horas
                hora_entrada = self.parsear_hora(entrada_str)
                hora_salida = self.parsear_hora(salida_str)
                
                if hora_entrada and hora_salida:
                    # Calcular diferencia
                    if hora_salida < hora_entrada:
                        # Asumir que salida es al día siguiente
                        hora_salida += timedelta(days=1)
                    
                    diferencia = hora_salida - hora_entrada
                    horas = diferencia.total_seconds() / 3600
                    
                    # Validar rango razonable (0-24 horas)
                    if 0 <= horas <= 24:
                        horas_calculadas.append(round(horas, 2))
                    else:
                        horas_calculadas.append(8.0)  # Default
                else:
                    horas_calculadas.append(8.0)  # Default
                    
            except Exception:
                horas_calculadas.append(8.0)  # Default
        
        return pd.Series(horas_calculadas)
    
    def parsear_hora(self, hora_str: str) -> Optional[datetime]:
        """Parsea string de hora a datetime"""
        patrones_hora = [
            r'(\d{1,2}):(\d{2}):(\d{2})',  # HH:MM:SS
            r'(\d{1,2}):(\d{2})',          # HH:MM
            r'(\d{3,4})',                   # HHMM
        ]
        
        for patron in patrones_hora:
            match = re.match(patron, hora_str)
            if match:
                try:
                    if len(match.groups()) == 3:  # HH:MM:SS
                        h, m, s = map(int, match.groups())
                        return datetime.now().replace(hour=h, minute=m, second=s, microsecond=0)
                    elif len(match.groups()) == 2:  # HH:MM
                        h, m = map(int, match.groups())
                        return datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
                    elif len(match.groups()) == 1:  # HHMM
                        hhmm = match.groups()[0]
                        if len(hhmm) == 3:
                            h, m = int(hhmm[0]), int(hhmm[1:3])
                        else:
                            h, m = int(hhmm[0:2]), int(hhmm[2:4])
                        return datetime.now().replace(hour=h, minute=m, second=0, microsecond=0)
                except ValueError:
                    continue
        
        return None
    
    def generar_reporte_excel(self, ruta_salida: str) -> bool:
        """
        Genera reporte profesional en Excel con formato avanzado
        
        Args:
            ruta_salida: Ruta donde guardar el archivo Excel
            
        Returns:
            bool: True si se generó correctamente
        """
        try:
            if self.resumen_mensual is None:
                self.errores.append("❌ No hay datos procesados para el reporte")
                return False
            
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remover hoja por defecto
            
            # Crear hojas
            self.crear_hoja_resumen_ejecutivo(wb)
            self.crear_hoja_detalle_mensual(wb)
            self.crear_hoja_analisis_empleados(wb)
            self.crear_hoja_estadisticas(wb)
            
            # Guardar archivo
            wb.save(ruta_salida)
            print(f"✅ Reporte Excel generado: {ruta_salida}")
            return True
            
        except Exception as e:
            self.errores.append(f"❌ Error al generar reporte Excel: {str(e)}")
            return False
    
    def crear_hoja_resumen_ejecutivo(self, wb: openpyxl.Workbook):
        """Crea hoja de resumen ejecutivo"""
        ws = wb.create_sheet("📊 Resumen Ejecutivo")
        
        # Título principal
        ws.merge_cells('A1:F3')
        titulo_cell = ws['A1']
        titulo_cell.value = "REPORTE DE HORAS TRABAJADAS - RESUMEN EJECUTIVO"
        titulo_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        titulo_cell.fill = PatternFill(start_color=self.COLORES_EXCEL['encabezado'], 
                                      end_color=self.COLORES_EXCEL['encabezado'], 
                                      fill_type='solid')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Información del reporte
        fila_actual = 5
        ws[f'A{fila_actual}'] = "Fecha de Generación:"
        ws[f'B{fila_actual}'] = datetime.now().strftime('%d de %B de %Y')
        ws[f'A{fila_actual+1}'] = "Total de Empleados:"
        ws[f'B{fila_actual+1}'] = len(self.resumen_mensual[self.resumen_mensual.columns[0]].unique())
        ws[f'A{fila_actual+2}'] = "Período Analizado:"
        ws[f'B{fila_actual+2}'] = f"{self.resumen_mensual['fecha_inicio'].min().strftime('%B %Y')} - {self.resumen_mensual['fecha_fin'].max().strftime('%B %Y')}"
        
        # Estadísticas generales
        fila_actual = 9
        ws.merge_cells(f'A{fila_actual}:F{fila_actual}')
        ws[f'A{fila_actual}'] = "ESTADÍSTICAS GENERALES"
        self.aplicar_estilo_encabezado(ws[f'A{fila_actual}'])
        
        fila_actual += 2
        estadisticas = self.calcular_estadisticas_generales()
        
        for stat_name, stat_value in estadisticas.items():
            ws[f'A{fila_actual}'] = stat_name
            ws[f'B{fila_actual}'] = stat_value
            fila_actual += 1
    
    def crear_hoja_detalle_mensual(self, wb: openpyxl.Workbook):
        """Crea hoja con detalle mensual por empleado"""
        ws = wb.create_sheet("📅 Detalle Mensual")
        
        # Configurar encabezados
        encabezados = [
            "Empleado", "Año", "Mes", "Total Horas", 
            "Días Trabajados", "Promedio Diario", "Fecha Inicio", "Fecha Fin"
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            self.aplicar_estilo_encabezado(cell)
        
        # Insertar datos
        for row, (_, datos) in enumerate(self.resumen_mensual.iterrows(), 2):
            ws.cell(row=row, column=1, value=datos[self.resumen_mensual.columns[0]])
            ws.cell(row=row, column=2, value=int(datos['año']))
            ws.cell(row=row, column=3, value=datos['nombre_mes'])
            
            # Horas totales con formato condicional
            cell_horas = ws.cell(row=row, column=4, value=float(datos['total_horas']))
            if datos['total_horas'] < 120:  # Menos de 120 horas/mes
                cell_horas.fill = PatternFill(start_color=self.COLORES_EXCEL['alerta'], 
                                            end_color=self.COLORES_EXCEL['alerta'], 
                                            fill_type='solid')
            elif datos['total_horas'] > 200:  # Más de 200 horas/mes
                cell_horas.fill = PatternFill(start_color=self.COLORES_EXCEL['info'], 
                                            end_color=self.COLORES_EXCEL['info'], 
                                            fill_type='solid')
            
            ws.cell(row=row, column=5, value=int(datos['dias_trabajados']))
            ws.cell(row=row, column=6, value=round(float(datos['promedio_diario']), 2))
            ws.cell(row=row, column=7, value=datos['fecha_inicio'].strftime('%d/%m/%Y'))
            ws.cell(row=row, column=8, value=datos['fecha_fin'].strftime('%d/%m/%Y'))
        
        # Autoajustar columnas
        self.autoajustar_columnas(ws)
        
        # Agregar filtros
        ws.auto_filter.ref = f"A1:{chr(65+len(encabezados)-1)}{len(self.resumen_mensual)+1}"
    
    def crear_hoja_analisis_empleados(self, wb: openpyxl.Workbook):
        """Crea análisis por empleado"""
        ws = wb.create_sheet("👥 Análisis por Empleado")
        
        # Agrupar por empleado
        analisis_empleados = self.resumen_mensual.groupby(self.resumen_mensual.columns[0]).agg({
            'total_horas': ['sum', 'mean', 'std', 'min', 'max'],
            'dias_trabajados': ['sum', 'mean'],
            'promedio_diario': 'mean'
        }).round(2)
        
        # Encabezados
        encabezados = [
            "Empleado", "Total Horas Acumuladas", "Promedio Mensual", 
            "Desviación Estándar", "Mes Mínimo", "Mes Máximo",
            "Total Días", "Promedio Días/Mes", "Horas Promedio/Día"
        ]
        
        for col, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col, value=encabezado)
            self.aplicar_estilo_encabezado(cell)
        
        # Insertar datos de análisis
        for row, (empleado, datos) in enumerate(analisis_empleados.iterrows(), 2):
            ws.cell(row=row, column=1, value=empleado)
            ws.cell(row=row, column=2, value=float(datos[('total_horas', 'sum')]))
            ws.cell(row=row, column=3, value=float(datos[('total_horas', 'mean')]))
            ws.cell(row=row, column=4, value=float(datos[('total_horas', 'std')]) if not pd.isna(datos[('total_horas', 'std')]) else 0)
            ws.cell(row=row, column=5, value=float(datos[('total_horas', 'min')]))
            ws.cell(row=row, column=6, value=float(datos[('total_horas', 'max')]))
            ws.cell(row=row, column=7, value=int(datos[('dias_trabajados', 'sum')]))
            ws.cell(row=row, column=8, value=float(datos[('dias_trabajados', 'mean')]))
            ws.cell(row=row, column=9, value=float(datos[('promedio_diario', 'mean')]))
        
        self.autoajustar_columnas(ws)
    
    def crear_hoja_estadisticas(self, wb: openpyxl.Workbook):
        """Crea hoja con gráficos y estadísticas avanzadas"""
        ws = wb.create_sheet("📈 Estadísticas y Gráficos")
        
        # Título
        ws.merge_cells('A1:H3')
        titulo_cell = ws['A1']
        titulo_cell.value = "ANÁLISIS ESTADÍSTICO DE HORAS TRABAJADAS"
        self.aplicar_estilo_encabezado(titulo_cell)
        
        # Estadísticas por mes
        fila_actual = 5
        ws[f'A{fila_actual}'] = "RESUMEN POR MES"
        self.aplicar_estilo_encabezado(ws[f'A{fila_actual}'])
        
        fila_actual += 2
        resumen_mensual_global = self.resumen_mensual.groupby(['año', 'nombre_mes']).agg({
            'total_horas': ['sum', 'mean', 'count']
        }).round(2)
        
        # Encabezados para resumen mensual
        encabezados_mes = ["Año", "Mes", "Total Horas", "Promedio por Empleado", "Empleados Activos"]
        for col, encabezado in enumerate(encabezados_mes, 1):
            ws.cell(row=fila_actual, column=col, value=encabezado).font = Font(bold=True)
        
        fila_actual += 1
        for (año, mes), datos in resumen_mensual_global.iterrows():
            ws.cell(row=fila_actual, column=1, value=int(año))
            ws.cell(row=fila_actual, column=2, value=mes)
            ws.cell(row=fila_actual, column=3, value=float(datos[('total_horas', 'sum')]))
            ws.cell(row=fila_actual, column=4, value=float(datos[('total_horas', 'mean')]))
            ws.cell(row=fila_actual, column=5, value=int(datos[('total_horas', 'count')]))
            fila_actual += 1
        
        self.autoajustar_columnas(ws)
    
    def aplicar_estilo_encabezado(self, cell):
        """Aplica estilo a celdas de encabezado"""
        cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=self.COLORES_EXCEL['encabezado'], 
                               end_color=self.COLORES_EXCEL['encabezado'], 
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def autoajustar_columnas(self, ws):
        """Autoajusta el ancho de las columnas"""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            
            # Find a non-merged cell to get the column letter
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            
            if column_letter is None:
                continue
            
            # Calculate max length
            for cell in column_cells:
                if not isinstance(cell, MergedCell) and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Set column width
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def calcular_estadisticas_generales(self) -> Dict[str, Union[int, float, str]]:
        """Calcula estadísticas generales del conjunto de datos"""
        stats = {}
        
        stats["Total Horas Registradas"] = f"{self.resumen_mensual['total_horas'].sum():,.2f} horas"
        stats["Promedio Mensual por Empleado"] = f"{self.resumen_mensual['total_horas'].mean():.2f} horas"
        stats["Empleado con Más Horas"] = self.resumen_mensual.loc[
            self.resumen_mensual['total_horas'].idxmax(), self.resumen_mensual.columns[0]
        ]
        stats["Máximo Horas en un Mes"] = f"{self.resumen_mensual['total_horas'].max():.2f} horas"
        stats["Mínimo Horas en un Mes"] = f"{self.resumen_mensual['total_horas'].min():.2f} horas"
        stats["Promedio de Días Trabajados"] = f"{self.resumen_mensual['dias_trabajados'].mean():.1f} días"
        
        return stats


class InterfazHR(tk.Tk if GUI_AVAILABLE else object):
    """
    Interfaz gráfica para la aplicación BioAdmin HR
    Solo disponible si tkinter está instalado
    """
    
    def __init__(self):
        if not GUI_AVAILABLE:
            raise ImportError("Interfaz gráfica no disponible: tkinter no instalado")
            
        super().__init__()
        
        self.procesador = BioadminHRProcessor()
        self.configurar_ventana()
        self.crear_widgets()
        
    def configurar_ventana(self):
        """Configura la ventana principal"""
        self.title("🏢 FORMAS - Sistema BioAdmin HR v2.0")
        self.geometry("800x600")
        self.configure(bg='#f0f0f0')
        
        # Centrar ventana
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (800 // 2)
        y = (self.winfo_screenheight() // 2) - (600 // 2)
        self.geometry(f"800x600+{x}+{y}")
        
        # Ícono (si existe)
        try:
            self.iconbitmap("formas_icon.ico")
        except:
            pass
    
    def crear_widgets(self):
        """Crea los widgets de la interfaz"""
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        titulo = ttk.Label(main_frame, text="Sistema de Procesamiento BioAdmin HR", 
                          font=('Arial', 16, 'bold'))
        titulo.pack(pady=(0, 20))
        
        # Frame de archivos
        archivo_frame = ttk.LabelFrame(main_frame, text="📁 Selección de Archivos", padding=10)
        archivo_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Botón para seleccionar archivo BioAdmin
        ttk.Button(archivo_frame, text="🔍 Seleccionar Archivo BioAdmin", 
                  command=self.seleccionar_archivo_bioadmin, 
                  width=30).pack(pady=5)
        
        self.label_archivo = ttk.Label(archivo_frame, text="Ningún archivo seleccionado", 
                                      foreground="gray")
        self.label_archivo.pack(pady=5)
        
        # Frame de procesamiento
        proceso_frame = ttk.LabelFrame(main_frame, text="⚙️ Procesamiento", padding=10)
        proceso_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(proceso_frame, text="🔄 Procesar Datos", 
                  command=self.procesar_datos, 
                  width=20).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(proceso_frame, text="📊 Generar Reporte Excel", 
                  command=self.generar_reporte, 
                  width=20).pack(side=tk.LEFT, padx=5)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.pack(fill=tk.X, pady=5)
        
        # Área de logs
        log_frame = ttk.LabelFrame(main_frame, text="📝 Registro de Actividad", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.text_log = ScrolledText(log_frame, height=15, wrap=tk.WORD)
        self.text_log.pack(fill=tk.BOTH, expand=True)
        
        # Frame de botones inferior
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(button_frame, text="🗑️ Limpiar Log", 
                  command=self.limpiar_log).pack(side=tk.LEFT)
        
        ttk.Button(button_frame, text="❓ Ayuda", 
                  command=self.mostrar_ayuda).pack(side=tk.RIGHT)
        
        # Log inicial
        self.log("🚀 Sistema BioAdmin HR iniciado correctamente")
        self.log("📋 Seleccione un archivo BioAdmin para comenzar")
        
        # Variables de estado
        self.archivo_seleccionado = None
    
    def log(self, mensaje: str):
        """Agrega mensaje al log con timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.text_log.insert(tk.END, f"[{timestamp}] {mensaje}\n")
        self.text_log.see(tk.END)
        self.update()
    
    def limpiar_log(self):
        """Limpia el área de log"""
        self.text_log.delete(1.0, tk.END)
        self.log("🧹 Log limpiado")
    
    def seleccionar_archivo_bioadmin(self):
        """Selecciona archivo BioAdmin"""
        tipos_archivo = [
            ("Archivos BioAdmin", "*.csv;*.xlsx;*.xls;*.txt"),
            ("Archivos CSV", "*.csv"),
            ("Archivos Excel", "*.xlsx;*.xls"),
            ("Archivos de texto", "*.txt"),
            ("Todos los archivos", "*.*")
        ]
        
        archivo = filedialog.askopenfilename(
            title="Seleccionar Archivo BioAdmin",
            filetypes=tipos_archivo,
            initialdir=os.path.expanduser("~")
        )
        
        if archivo:
            self.archivo_seleccionado = archivo
            nombre_archivo = os.path.basename(archivo)
            self.label_archivo.config(text=f"✅ {nombre_archivo}", foreground="green")
            self.log(f"📁 Archivo seleccionado: {nombre_archivo}")
        else:
            self.log("⚠️ Selección de archivo cancelada")
    
    def procesar_datos(self):
        """Procesa los datos del archivo seleccionado"""
        if not self.archivo_seleccionado:
            messagebox.showwarning("Advertencia", "Por favor seleccione un archivo BioAdmin primero")
            self.log("⚠️ No hay archivo seleccionado para procesar")
            return
        
        self.progress.start()
        self.log("🔄 Iniciando procesamiento de datos...")
        
        try:
            # Leer archivo BioAdmin
            self.log("📖 Leyendo archivo BioAdmin...")
            if not self.procesador.leer_archivo_bioadmin(self.archivo_seleccionado):
                self.mostrar_errores()
                return
            
            # Procesar horas trabajadas
            self.log("⚙️ Calculando horas trabajadas...")
            if not self.procesador.procesar_horas_trabajadas():
                self.mostrar_errores()
                return
            
            # Mostrar estadísticas básicas
            if self.procesador.resumen_mensual is not None:
                total_empleados = len(self.procesador.resumen_mensual[
                    self.procesador.resumen_mensual.columns[0]].unique())
                total_registros = len(self.procesador.resumen_mensual)
                
                self.log(f"✅ Procesamiento completado exitosamente")
                self.log(f"👥 Empleados procesados: {total_empleados}")
                self.log(f"📊 Registros mensuales: {total_registros}")
                self.log("🎯 Datos listos para generar reporte")
                
                messagebox.showinfo("Éxito", 
                                  f"Datos procesados correctamente!\n\n"
                                  f"• Empleados: {total_empleados}\n"
                                  f"• Registros mensuales: {total_registros}")
            
        except Exception as e:
            self.log(f"❌ Error durante el procesamiento: {str(e)}")
            messagebox.showerror("Error", f"Error durante el procesamiento:\n{str(e)}")
        
        finally:
            self.progress.stop()
    
    def generar_reporte(self):
        """Genera el reporte Excel"""
        if self.procesador.resumen_mensual is None:
            messagebox.showwarning("Advertencia", "Debe procesar los datos primero")
            self.log("⚠️ No hay datos procesados para generar reporte")
            return
        
        # Seleccionar ubicación de guardado
        archivo_salida = filedialog.asksaveasfilename(
            title="Guardar Reporte Excel",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfilename=f"Reporte_HR_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        
        if not archivo_salida:
            self.log("⚠️ Generación de reporte cancelada")
            return
        
        self.progress.start()
        self.log("📊 Generando reporte Excel profesional...")
        
        try:
            if self.procesador.generar_reporte_excel(archivo_salida):
                self.log(f"✅ Reporte generado exitosamente: {os.path.basename(archivo_salida)}")
                messagebox.showinfo("Éxito", 
                                  f"Reporte Excel generado correctamente!\n\n"
                                  f"Ubicación: {archivo_salida}")
                
                # Preguntar si desea abrir el archivo
                if messagebox.askyesno("Abrir Archivo", "¿Desea abrir el reporte generado?"):
                    try:
                        os.startfile(archivo_salida)  # Windows
                    except AttributeError:
                        try:
                            os.system(f"open {archivo_salida}")  # macOS
                        except:
                            os.system(f"xdg-open {archivo_salida}")  # Linux
            else:
                self.mostrar_errores()
                
        except Exception as e:
            self.log(f"❌ Error al generar reporte: {str(e)}")
            messagebox.showerror("Error", f"Error al generar reporte:\n{str(e)}")
        
        finally:
            self.progress.stop()
    
    def mostrar_errores(self):
        """Muestra errores acumulados"""
        if self.procesador.errores:
            errores_texto = "\n".join(self.procesador.errores)
            self.log("❌ Errores encontrados:")
            for error in self.procesador.errores:
                self.log(f"   {error}")
            
            messagebox.showerror("Errores Encontrados", errores_texto)
            self.procesador.errores.clear()
    
    def mostrar_ayuda(self):
        """Muestra ventana de ayuda"""
        ayuda_window = tk.Toplevel(self)
        ayuda_window.title("📖 Ayuda - Sistema BioAdmin HR")
        ayuda_window.geometry("600x500")
        ayuda_window.configure(bg='white')
        
        # Centrar ventana de ayuda
        ayuda_window.update_idletasks()
        x = (ayuda_window.winfo_screenwidth() // 2) - (600 // 2)
        y = (ayuda_window.winfo_screenheight() // 2) - (500 // 2)
        ayuda_window.geometry(f"600x500+{x}+{y}")
        
        # Contenido de ayuda
        ayuda_text = ScrolledText(ayuda_window, wrap=tk.WORD, padx=10, pady=10)
        ayuda_text.pack(fill=tk.BOTH, expand=True)
        
        contenido_ayuda = """
📖 SISTEMA BIOADMIN HR - GUÍA DE USO

🎯 PROPÓSITO
Este sistema procesa archivos de BioAdmin y genera reportes profesionales 
en Excel con análisis detallado de horas trabajadas por empleado.

📁 ARCHIVOS SOPORTADOS
• CSV (.csv) - Separados por comas, punto y coma o tabulación
• Excel (.xlsx, .xls) - Archivos de Microsoft Excel
• Texto (.txt) - Archivos de texto delimitados

📋 ESTRUCTURA ESPERADA DEL ARCHIVO
El archivo debe contener al menos estas columnas:
• Empleado/Nombre/Usuario - Identificación del empleado
• Fecha - Fecha del registro
• Hora Entrada/Inicio - Hora de llegada (opcional)
• Hora Salida/Fin - Hora de salida (opcional)

📊 FUNCIONALIDADES
• Cálculo automático de horas trabajadas por día
• Agrupación de datos por mes y empleado  
• Análisis estadístico completo
• Reportes Excel profesionales con:
  - Resumen ejecutivo
  - Detalle mensual por empleado
  - Análisis estadístico por empleado
  - Gráficos y métricas avanzadas

🔧 PROCESO DE USO
1. Seleccione un archivo BioAdmin usando "Seleccionar Archivo"
2. Presione "Procesar Datos" para analizar la información
3. Use "Generar Reporte Excel" para crear el informe final

⚠️ NOTAS IMPORTANTES
• El sistema maneja fechas en español automáticamente
• Si no hay horas de entrada/salida, asume 8 horas por día
• Los archivos deben tener encabezados en la primera fila
• Se valida la estructura del archivo automáticamente

🆘 SOPORTE TÉCNICO
Para soporte técnico, contacte al departamento de IT de FORMAS.

© 2024 Grupo Formas - Sistema BioAdmin HR v2.0
        """
        
        ayuda_text.insert(tk.END, contenido_ayuda)
        ayuda_text.config(state=tk.DISABLED)


def main():
    """Función principal de la aplicación"""
    try:
        # Verificar que los módulos necesarios estén instalados
        modulos_requeridos = ['pandas', 'openpyxl', 'numpy']
        modulos_faltantes = []
        
        for modulo in modulos_requeridos:
            try:
                __import__(modulo)
            except ImportError:
                modulos_faltantes.append(modulo)
        
        if modulos_faltantes:
            print(f"❌ Módulos faltantes: {', '.join(modulos_faltantes)}")
            print("💡 Instale con: pip install pandas openpyxl numpy")
            input("Presione Enter para salir...")
            return
        
        # Solo iniciar GUI si está disponible
        if not GUI_AVAILABLE:
            print("❌ Interfaz gráfica no disponible (tkinter no instalado)")
            print("💡 Use la versión de consola: python bioadmin_hr_console.py")
            return
            
        # Iniciar interfaz gráfica
        app = InterfazHR()
        
        # Configurar manejo de cierre
        def on_closing():
            if messagebox.askokcancel("Salir", "¿Está seguro que desea salir?"):
                app.destroy()
        
        app.protocol("WM_DELETE_WINDOW", on_closing)
        
        # Ejecutar aplicación
        app.mainloop()
        
    except Exception as e:
        print(f"❌ Error crítico al iniciar la aplicación: {e}")
        input("Presione Enter para salir...")


if __name__ == "__main__":
    print("🏢 Iniciando Sistema BioAdmin HR de FORMAS...")
    print("=" * 50)
    main()